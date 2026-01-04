import os
import sys
import django
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'edu.settings')
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
django.setup()

from applications.academico.models import Carrera, PeriodoAcademico

# Crear archivo Excel con datos de ejemplo
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Asignaturas"

# Estilos
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

# Encabezados
headers = ["Carrera", "Código", "Materia", "Semestre", "Créditos", "Descripción", "Prerrequisitos"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Datos de ejemplo - cadena de asignaturas con prerrequisitos
data = [
    ["Ingeniería en Sistemas", "ING-SIS-CAL1", "Cálculo I", 1, 4, "Fundamentos de cálculo", "-"],
    ["Ingeniería en Sistemas", "ING-SIS-ALG", "Álgebra", 1, 4, "Fundamentos de álgebra", "-"],
    ["Ingeniería en Sistemas", "ING-SIS-CAL2", "Cálculo II", 2, 4, "Cálculo avanzado", "ING-SIS-CAL1"],
    ["Ingeniería en Sistemas", "ING-SIS-ALG2", "Álgebra II", 2, 4, "Álgebra avanzada", "ING-SIS-ALG"],
    ["Ingeniería en Sistemas", "ING-SIS-ED", "Estructuras de Datos", 2, 4, "Estructuras de datos", "ING-SIS-ALG"],
    ["Ingeniería en Sistemas", "ING-SIS-POO", "Programación OO", 3, 5, "Programación orientada a objetos", "ING-SIS-ALG, ING-SIS-ED"],
    ["Ingeniería en Sistemas", "ING-SIS-BD1", "Bases de Datos I", 3, 4, "Intro a BD", "ING-SIS-ED"],
    ["Ingeniería en Sistemas", "ING-SIS-ARQ", "Arquitectura de Computadores", 3, 4, "Arquitectura", "ING-SIS-ALG"],
    ["Ingeniería en Sistemas", "ING-SIS-BD2", "Bases de Datos II", 4, 4, "BD avanzado", "ING-SIS-BD1"],
    ["Ingeniería en Sistemas", "ING-SIS-SO", "Sistemas Operativos", 4, 4, "Sistemas operativos", "ING-SIS-ARQ"],
]

for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Ajustar ancho de columnas
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 30
ws.column_dimensions['G'].width = 30

# Guardar
output_path = os.path.join(os.path.dirname(__file__), '..', '..', '..', 'asignaturas_con_prerrequisitos.xlsx')
wb.save(output_path)

print(f"✓ Archivo creado: {output_path}")
print(f"\nContenido:")
print("- Ingeniería en Sistemas con 10 asignaturas (semestres 1-4)")
print("- Asignaturas con cadenas de prerrequisitos")
print("- Ejemplo: POO requiere ALG + ED")
print("\nPuedes importar este archivo desde la interfaz de Asignaturas")
